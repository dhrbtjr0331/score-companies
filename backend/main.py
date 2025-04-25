# main.py
from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse  # Add this import
from jose import JWTError, jwt
from passlib.context import CryptContext
from typing import Optional, List, Dict, Any
from pydantic import BaseModel, Field
from datetime import datetime, timedelta
import os
import traceback
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship, Session
import dropbox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import tempfile
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# FastAPI app
app = FastAPI(title="Investment Scorecard API")

# CORS setup - Fix the CORS issue
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Your React app's URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# JWT Configuration
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Token URL
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# Database setup
SQLALCHEMY_DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./scorecard.db")
engine = create_engine(SQLALCHEMY_DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Dropbox configuration
DROPBOX_ACCESS_TOKEN = os.getenv("DROPBOX_ACCESS_TOKEN")

# Database Models
class User(Base):
    __tablename__ = "users"
    
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    first_name = Column(String)
    last_name = Column(String)
    hashed_password = Column(String)
    
    scorecards = relationship("Scorecard", back_populates="scored_by")

class Scorecard(Base):
    __tablename__ = "scorecards"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date)
    company_name = Column(String, index=True)
    sector = Column(String, index=True)
    investment_stage = Column(String)
    alignment = Column(Integer)
    team = Column(Integer)
    market = Column(Integer)
    product = Column(Integer)
    potential_return = Column(Integer)
    bold_excitement = Column(Integer)
    score = Column(Float)
    
    user_id = Column(Integer, ForeignKey("users.id"))
    scored_by = relationship("User", back_populates="scorecards")

# Create tables
Base.metadata.create_all(bind=engine)

# Dependency to get DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Pydantic Models (Request/Response models)
class Token(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"

class TokenData(BaseModel):
    username: Optional[str] = None

class UserBase(BaseModel):
    username: str

class UserCreate(UserBase):
    password: str
    retypePassword: str
    firstname: str
    lastname: str

class UserLogin(UserBase):
    password: str

class UserResponse(UserBase):
    first_name: str
    last_name: str
    
    class Config:
        orm_mode = True

class AuthStatus(BaseModel):
    isAuthenticated: bool
    user: Optional[UserResponse] = None

class ScoreBase(BaseModel):
    date: str
    company_name: str
    sector: str
    investment_stage: str
    alignment: int = Field(..., ge=0, le=10)
    team: int = Field(..., ge=0, le=10)
    market: int = Field(..., ge=0, le=10)
    product: int = Field(..., ge=0, le=10)
    potential_return: int = Field(..., ge=0, le=10)
    bold_excitement: int = Field(..., ge=0, le=10)

# Modified ScoreResponse to handle nested scores correctly
class ScoreResponse(BaseModel):
    id: int
    date: str
    company_name: str
    sector: str
    investment_stage: str
    alignment: int
    team: int
    market: int
    product: int
    potential_return: int
    bold_excitement: int
    score: float
    scored_by: Dict[str, str]
    scores: Dict[str, int]
    
    class Config:
        orm_mode = True

class RefreshToken(BaseModel):
    refresh_token: str

# Helper functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def get_user(db: Session, username: str):
    return db.query(User).filter(User.username == username).first()

def authenticate_user(db: Session, username: str, password: str):
    user = get_user(db, username)
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    return user

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def generate_jwt_tokens(user):
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    refresh_token = create_refresh_token(data={"sub": user.username})
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer"
    }

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    user = get_user(db, username=token_data.username)
    if user is None:
        raise credentials_exception
    return user

def validate_registration_fields(username, password, retype_password, firstname, lastname):
    if not username:
        return "Username is required."
    if not password:
        return "Password is required."
    if not retype_password:
        return "Retype password is required."
    if not firstname:
        return "First name is required."
    if not lastname:
        return "Last name is required."
    return None

def validate_scorecard_fields(date, company_name, sector, investment_stage, alignment, team, market, product, potential_return, bold_excitement):
    if not date:
        return "Date is required."
    if not company_name:
        return "Company Name is required."
    if not sector:
        return "Sector is required."
    if not investment_stage:
        return "Investment Stage is required."
    if not alignment:
        return "Alignment is required."
    if not team:
        return "Team is required."
    if not market:
        return "Market is required."
    if not product:
        return "Product is required."
    if not potential_return:
        return "Potential_return is required."
    if not bold_excitement:
        return "Bold_excitement is required."
    return None

def get_score(alignment, team, market, product, potential_return, bold_excitement):
    score = (alignment + market + product + bold_excitement) * (team + potential_return) / 80.00
    return round(score, 2)

# Excel and Dropbox Functions
def fetch_excel_from_dropbox(dropbox_path, local_path):
    """Fetches an Excel file from Dropbox and saves it locally."""
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)
    print(f"Token being used: '{DROPBOX_ACCESS_TOKEN}'")
    with open(local_path, "wb") as f:
        metadata, res = dbx.files_download(path=dropbox_path)
        f.write(res.content)
    print(f"File successfully fetched from Dropbox to: {local_path}")
    return local_path

def append_and_format_to_excel(file_path, scorecard):
    """Appends a new scorecard row to the Excel file with grouped companies and their average score."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    name = f"{scorecard['scored_by']['first_name']} {scorecard['scored_by']['last_name']}"
    
    # Create the new row
    new_row = [
        name,
        scorecard["company_name"],
        scorecard["date"],
        scorecard["sector"],
        scorecard["investment_stage"],
        scorecard["scores"]["alignment"],
        scorecard["scores"]["team"],
        scorecard["scores"]["market"],
        scorecard["scores"]["product"],
        scorecard["scores"]["potential_return"],
        scorecard["scores"]["bold_excitement"],
        scorecard["score"]
    ]

    # Append the new row
    sheet.append(new_row)

    # Read all rows
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    # Group rows by company name
    grouped_data = {}
    for row in data:
        company_name = row[1]  # Assuming Company Name is column B
        if company_name:
            grouped_data.setdefault(company_name, []).append(row)

    # Clear rows below the header
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    # Define a border style for the "Total Score" section
    border_style_1 = Border(
        left=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    border_style_2 = Border(
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )

    # Write grouped data back with formatting
    for company, rows in grouped_data.items():
        scoreSum = 0
        count = 0
        if rows:  # Only process non-empty groups
            for row in rows:
                sheet.append(row)
                scoreSum += row[11]  # Assuming 'Score' column is at index 11
                count += 1

            # Calculate average score
            avg_score = round(scoreSum / count, 2) if count > 0 else 0

            # Get the row index for "Total Score"
            total_score_row = sheet.max_row + 1
            score_col = len(header)  # Last column (score column)
            label_col = score_col - 1  # Column before score

            # Merge cells for "Total Score"
            sheet.merge_cells(start_row=total_score_row, start_column=label_col, 
                              end_row=total_score_row, end_column=score_col - 1)

            # Set "Total Score" text
            total_score_cell = sheet.cell(row=total_score_row, column=label_col)
            total_score_cell.value = "Total Score"
            total_score_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_score_cell.border = border_style_1  # Apply border

            # Set total score value in the last column
            score_cell = sheet.cell(row=total_score_row, column=score_col)
            score_cell.value = avg_score
            score_cell.alignment = Alignment(horizontal="center", vertical="center")
            score_cell.border = border_style_2  # Apply border

            # Append two blank rows after each company section
            sheet.append([""] * len(header))

    # Remove trailing blank rows
    while all(cell.value is None for cell in sheet[sheet.max_row]):
        sheet.delete_rows(sheet.max_row)

    # Format the header row
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(file_path)
    workbook.close()
    print(f"Excel file updated and saved locally at: {file_path}")

def upload_excel_to_dropbox(local_path, dropbox_path):
    """Uploads an Excel file to Dropbox."""
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)
    with open(local_path, "rb") as f:
        res = dbx.files_upload(f.read(), path=dropbox_path, mode=dropbox.files.WriteMode.overwrite)
    print(f"File successfully uploaded to Dropbox: {dropbox_path}")
    return f"Updated file uploaded to Dropbox at {dropbox_path}"

# API Routes
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    print(f"Unhandled exception: {exc}")
    print(traceback.format_exc())
    if isinstance(exc, HTTPException):
        return JSONResponse(
            status_code=exc.status_code,
            content={"error": exc.detail}
        )
    return JSONResponse(
        status_code=500,
        content={"error": f"Internal server error: {str(exc)}"}
    )

# Add OPTIONS handler for preflight requests
@app.options("/api/score-company/")
async def options_score_company():
    return JSONResponse(content={}, headers={
        "Access-Control-Allow-Origin": "http://localhost:3000",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "*"
    })

@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    tokens = generate_jwt_tokens(user)
    return tokens

@app.post("/api/login/", response_model=Token)
async def login(user_data: UserLogin, db: Session = Depends(get_db)):
    user = authenticate_user(db, user_data.username, user_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
        )
    tokens = generate_jwt_tokens(user)
    return tokens

@app.post("/api/register/", response_model=Dict[str, Any])
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    # Validate fields
    validation_error = validate_registration_fields(
        user_data.username, 
        user_data.password, 
        user_data.retypePassword, 
        user_data.firstname, 
        user_data.lastname
    )
    
    if validation_error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=validation_error)

    if user_data.password != user_data.retypePassword:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Passwords do not match.")

    if get_user(db, user_data.username):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Username already exists.")

    hashed_password = get_password_hash(user_data.password)
    db_user = User(
        username=user_data.username,
        hashed_password=hashed_password,
        first_name=user_data.firstname,
        last_name=user_data.lastname
    )
    
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    tokens = generate_jwt_tokens(db_user)
    return {
        "message": "Registration successful.",
        **tokens
    }

@app.post("/api/refresh-token/", response_model=Token)
async def refresh_token(refresh_token_data: RefreshToken, db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(refresh_token_data.refresh_token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid refresh token",
            )
        
        user = get_user(db, username)
        if user is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="User not found",
            )
            
        tokens = generate_jwt_tokens(user)
        return tokens
        
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid refresh token",
        )

@app.get("/api/is-authenticated/", response_model=AuthStatus)
async def is_authenticated(current_user: User = Depends(get_current_user)):
    return {
        "isAuthenticated": True,
        "user": {
            "username": current_user.username,
            "first_name": current_user.first_name,
            "last_name": current_user.last_name
        }
    }

@app.post("/api/logout/")
async def logout(current_user: User = Depends(get_current_user)):
    # JWT tokens are stateless, so logout is handled client-side
    return {
        "message": "Logout successful. Please ensure that tokens are cleared on the client side to complete the logout process."
    }

@app.post("/api/score-company/", response_model=ScoreResponse)
async def score_company(
    score_data: ScoreBase, 
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        print(f"Received score data: {score_data}")
        
        # Calculate score
        score = get_score(
            score_data.alignment,
            score_data.team,
            score_data.market,
            score_data.product,
            score_data.potential_return,
            score_data.bold_excitement
        )
        
        # Create scorecard data for Excel
        scorecard = {
            "id": 1,  # Mock ID for Excel processing
            "date": score_data.date,
            "company_name": score_data.company_name,
            "sector": score_data.sector,
            "investment_stage": score_data.investment_stage,
            "scores": {
                "alignment": score_data.alignment,
                "team": score_data.team,
                "market": score_data.market,
                "product": score_data.product,
                "potential_return": score_data.potential_return,
                "bold_excitement": score_data.bold_excitement,
            },
            "score": score,
            "scored_by": {
                "first_name": current_user.first_name,
                "last_name": current_user.last_name,
            }
        }

        # Save to database
        try:
            db_scorecard = Scorecard(
                date=datetime.strptime(score_data.date, "%Y-%m-%d").date(),
                company_name=score_data.company_name,
                sector=score_data.sector,
                investment_stage=score_data.investment_stage,
                alignment=score_data.alignment,
                team=score_data.team,
                market=score_data.market,
                product=score_data.product,
                potential_return=score_data.potential_return,
                bold_excitement=score_data.bold_excitement,
                score=score,
                user_id=current_user.id
            )
            
            db.add(db_scorecard)
            db.commit()
            db.refresh(db_scorecard)
            
            # Update scorecard with database ID
            scorecard["id"] = db_scorecard.id
            print("Successfully saved to database")
        except Exception as db_error:
            print(f"Database error: {str(db_error)}")
            print(traceback.format_exc())
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Database error: {str(db_error)}"
            )

        # Handle Excel export to Dropbox if configured
        if DROPBOX_ACCESS_TOKEN and DROPBOX_ACCESS_TOKEN != 'your-dropbox-access-token-here':
            # Define paths for Dropbox
            dropbox_path = "/scorecards.xlsx"
            local_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name

            try:
                # Step 1: Fetch file from Dropbox
                print(f"Fetching file from Dropbox: {dropbox_path}")
                fetch_excel_from_dropbox(dropbox_path, local_path)
                print(f"File successfully fetched from Dropbox: {local_path}")

                # Step 2: Update Excel file
                print("Appending and formatting the Excel file...")
                append_and_format_to_excel(local_path, scorecard)
                print(f"Excel file updated at: {local_path}")

                # Ensure file exists before upload
                if not os.path.exists(local_path):
                    print(f"File does not exist at {local_path}, aborting upload.")
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
                        detail="Local file was not created properly"
                    )

                # Step 3: Upload file back to Dropbox
                print(f"Uploading updated file to Dropbox: {dropbox_path}")
                message = upload_excel_to_dropbox(local_path, dropbox_path)
                print(message)
            except Exception as dropbox_error:
                print(f"Dropbox operation failed: {str(dropbox_error)}")
                print(traceback.format_exc())
                # Don't fail the entire operation if Dropbox fails
                print("Continuing without Dropbox sync due to error")
            finally:
                # Clean up the local temporary file
                if os.path.exists(local_path):
                    os.unlink(local_path)
                    print(f"Temporary file cleaned up: {local_path}")
        else:
            print("Dropbox token not configured. Skipping Excel export.")

        # Return full scorecard response with individual scores at the top level
        return {
            "id": scorecard["id"],
            "date": scorecard["date"],
            "company_name": scorecard["company_name"],
            "sector": scorecard["sector"],
            "investment_stage": scorecard["investment_stage"],
            # Individual scores at top level to match ScoreResponse model
            "alignment": score_data.alignment,
            "team": score_data.team,
            "market": score_data.market,
            "product": score_data.product,
            "potential_return": score_data.potential_return,
            "bold_excitement": score_data.bold_excitement,
            "score": scorecard["score"],
            "scored_by": scorecard["scored_by"],
            # Also keep scores dict for backward compatibility if needed
            "scores": scorecard["scores"]
        }
    except Exception as e:
        print(f"Error in score_company endpoint: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Server error: {str(e)}"
        )

@app.get("/api/sectors/", response_model=List[str])
async def get_sectors(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sectors = db.query(Scorecard.sector).distinct().all()
    return [sector[0] for sector in sectors if sector[0]]

@app.get("/api/company-names/", response_model=List[str])
async def get_company_names(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    company_names = db.query(Scorecard.company_name).distinct().all()
    return [name[0] for name in company_names if name[0]]

@app.get("/api/scored-companies/", response_model=List[ScoreResponse])
async def get_scored_companies(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    scorecards = db.query(Scorecard).join(User).all()
    
    response_data = []
    for scorecard in scorecards:
        response_data.append({
            "id": scorecard.id,
            "date": str(scorecard.date),
            "company_name": scorecard.company_name,
            "sector": scorecard.sector,
            "investment_stage": scorecard.investment_stage,
            # Individual scores at top level to match ScoreResponse model
            "alignment": scorecard.alignment,
            "team": scorecard.team,
            "market": scorecard.market,
            "product": scorecard.product,
            "potential_return": scorecard.potential_return,
            "bold_excitement": scorecard.bold_excitement,
            "score": scorecard.score,
            "scored_by": {
                "first_name": scorecard.scored_by.first_name if scorecard.scored_by else None,
                "last_name": scorecard.scored_by.last_name if scorecard.scored_by else None,
            },
            # Also keep scores dict for backward compatibility if needed
            "scores": {
                "alignment": scorecard.alignment,
                "team": scorecard.team,
                "market": scorecard.market,
                "product": scorecard.product,
                "potential_return": scorecard.potential_return,
                "bold_excitement": scorecard.bold_excitement,
            }
        })
    return response_data

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)